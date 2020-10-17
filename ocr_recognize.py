#!/usr/bin/python3
import time
import pytesseract
import sys, re, os
from openpyxl import load_workbook
from PIL import Image

basePath = "/home/pi/Scripts/"

def init():
    global img 
    filePath = os.path.join(basePath, "recognize.png")
    img = Image.open (filePath)

def parseTeseractResults(stringToParse, myDictionary):  
    for line in stringToParse.splitlines():
        items = line.split()
        #dict_obj[key] = [dict_obj[key]]
        if "level" not in myDictionary:
            myDictionary["level"]     = [items[0]]
            myDictionary["page_num"]  = [items[1]]
            myDictionary["block_num"] = [items[2]]     
            myDictionary["par_num"]   = [items[3]]
            myDictionary["line_num"]  = [items[4]]
            myDictionary["word_num"]  = [items[5]]
            myDictionary["left"]      = [items[6]]
            myDictionary["top"]       = [items[7]]
            myDictionary["width"]     = [items[8]]
            myDictionary["height"]    = [items[9]]
            myDictionary["conf"]      = [items[10]] 
            try:
                myDictionary["text"]      = [items[11]] 
            except: 
                myDictionary["text"]      = "NA" 
        else: 
            myDictionary["level"].append(items[0])
            myDictionary["page_num"].append(items[1])
            myDictionary["block_num"].append(items[2])     
            myDictionary["par_num"].append(items[3])
            myDictionary["line_num"].append(items[4])
            myDictionary["word_num"].append(items[5])
            myDictionary["left"].append(items[6])
            myDictionary["top"].append(items[7])
            myDictionary["width"].append(items[8])
            myDictionary["height"].append(items[9])
            myDictionary["conf"].append(items[10]) 
            try:
                myDictionary["text"].append(items[11])

            except:
                myDictionary["text"].append("NA") 

# use teseract to recognize digits 
def recognize():
    custom_config = r"-c tessedit_char_whitelist=0123456789 --psm 6 -l digits"
    text = pytesseract.image_to_data(img, config=custom_config)
    print (text)
    parsedData = {}
    parseTeseractResults(text, parsedData)
    return parsedData

# decide if recognized data make sense
def getValidityAndConsumption(myDict, previousValidValues):
    consumption = ""
    index = 0
    recognized_digits = 0
    expected_digits = 8

    for item in myDict["text"]:
        if item == "NA" or item == "text":
            pass
        else:
            recognized_digits += 1 
            if (float(myDict["conf"][index]) < 70.0):
                return [2, -1]
            consumption = consumption + item 

        index += 1

    if recognized_digits != expected_digits :
        return [1,-1]
    consumption = float (consumption)/1000
    return [0, consumption]

def getDateTime():
    result = re.search(r'\/([0-9|_]+)\.\w+$', str(sys.argv[1]))
    fullDateTime = str(result.group(1))
    fullDateTime = re.split('_', fullDateTime)
    global date, actTime
    date    = str(fullDateTime[2] + '.' + fullDateTime[1] + '.' + fullDateTime[0])
    actTime = str(fullDateTime[3] + ':' + fullDateTime[4])

# To write consumtion and 
def writeIntoTxt (file, actualConsumption):
    file = open(file, 'a')  
    file.write('\n'+ date +"\t" + actTime + "\t" + str(actualConsumption))
    file.close()   

def writeIntoXlsx(actualConsumption):
    # writing into xlsx
    wb = load_workbook('/home/pi/Desktop/technicka.xlsx')
    ws = wb['Sheet']


    row_count = ws.max_row
    #column_count = ws.max_column
    #writeTo = 'C' + str(row_count+1)

    # Write actual consumtion
    ws[('C' + str(row_count+1))] = str(actualConsumption)
    ws[('A' + str(row_count+1))] = date
    ws[('B' + str(row_count+1))] = actTime

    text = ws['A1'].value
    print (text)
    wb.save('/home/pi/Desktop/technicka.xlsx')


init()
getDateTime()
# try several times to get better picture (the wheel on the flowmeter could be moving)
for i in range(6):
    consumptionDictionary = recognize()
    retVal = getValidityAndConsumption(consumptionDictionary, [0,0])
    consumption = retVal[1]
    if consumption != -1:
        break

    time.sleep(5.0)
    print ("another try")

print ("Actual Consumption:", consumption )
if consumption != -1:
    writeIntoTxt(os.path.join(basePath, "web/text.txt"), consumption)
    writeIntoXlsx(consumption)